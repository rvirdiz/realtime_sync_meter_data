#!/usr/bin/env python
import MySQLdb
import urllib2, base64, xmltodict
import dateutil.parser as dateparser
from dateutil.relativedelta import relativedelta
import requests
import openpyxl
import sys
from json import dumps
from requests.auth import HTTPBasicAuth
from datetime import datetime
from datetime import timedelta
from openpyxl.styles import PatternFill, Alignment, Font
from urlparse import urljoin as combine_url


"""Default Variables"""
### stg db default ###
mysql_url = "localhost"
mysql_user = "username"
mysql_password = "password"
mysql_db = "building_consumption"

### Energystar ###
energy_default_url = "https://portfoliomanager.energystar.gov/ws/"
energy_username = 'username'
energy_password = 'password'

### Leedon values ###
performance_default_url = "http://plaque.sbx4.leedon.io"
performance_username = 'username'
performance_password = 'password'

### Meters List Table Name ###
energystar_meters = "meters"
### Meter Consumption Data Table Name ###
energystar_data = "meter_consumption"
### Property Table Name used to get only list of LeedId ###
energystar_property = "properties"

### LEEDON BASE URL ###
leedon_url = performance_default_url

### LEEDON AUTHENTICATION ###
leed_auth = HTTPBasicAuth(performance_username,performance_password)
### ENERGYSTAR AUTHENTICATION ###
energy_auth = base64.encodestring('%s:%s' %(energy_username,energy_password)).replace('\n', '')

### ENERGYSTAR BASE URL ###
energystar_url = energy_default_url

### Excel Filename XLSX format ###
leedon_excel = "meter_consumption.xlsx"



"""Function for connecting to the db"""
def connect_db():
	try:
		db = MySQLdb.connect(host=mysql_url,user=mysql_user, passwd=mysql_password,db=mysql_db)
		cursor = db.cursor()
	except Exception as e:
		print "---Exception in connect_db()---", e

	return db, cursor


"""Timestamp for calculating the running time"""
FMT = '%H:%M:%S'
FMD = '%Y-%m-%d'
FMF = '%Y-%m-%d %H:%M:%S'

"""Excel Sheet Styles"""
font = Font(name='Calibri', size=11, bold=True, color='FFFFFFFF')
fill = PatternFill(fill_type="lightGray", start_color='FFFFFFFF', end_color='FF000000')
alignment = Alignment(horizontal='center', vertical='center', text_rotation=0, wrap_text=False, shrink_to_fit=False,\
 indent=2)

"""Excel sheet to save the Overlapping data while posting to leedon and check if it exists or not"""
try:
	wb = openpyxl.load_workbook(leedon_excel)
	ws1 = wb.get_sheet_by_name('Overlapping Meter Data')
	ws2 = wb.get_sheet_by_name('Updated Overlapping Meter Data')
except:
	wb = openpyxl.Workbook()
	ws1 = wb.active
	ws1.title = "Overlapping Meter Data"
	ws1.append(["Project Id", "Meter Name", "Actual Start Date", "Actual End Date", "Usage"])

	ws2 = wb.create_sheet(title="Updated Overlapping Meter Data")
	ws2.append(["Project Id", "Meter Name", "Actual Start Date", "Actual End Date", "Usage", "Modified Start Date",\
	 "Modified End Date"])

	for row in ws1.iter_rows('A1:E1'):
		for cell in row:
			cell.font = font
			cell.alignment = alignment
			cell.fill = fill

	for row in ws2.iter_rows('A1:G1'):
		for cell in row:
			cell.font = font
			cell.alignment = alignment
			cell.fill = fill

	sheet_columns = ["A", "B", "C", "D", "E"]
	for i in sheet_columns:
		ws1.column_dimensions[i].width = 20
		ws1.column_dimensions[i].alignment = alignment

	sheet_columns.append("F")
	sheet_columns.append("G")

	for i in sheet_columns:
		ws2.column_dimensions[i].width = 20
		ws2.column_dimensions[i].alignment = alignment

	ws1.column_dimensions["B"].width = 35
	ws2.column_dimensions["B"].width = 35

 	wb.save(filename = leedon_excel)

#######################################################################################################################

"""urljoin function defined to join the base_url and the meter consumption link url"""
def urljoin(*args):
    """
    Joins given arguments into a url. Trailing but not leading slashes are
    stripped for each argument.
    """

    return "".join(map(lambda x: str(x).rstrip('/'), args))

#######################################################################################################################

"""Get the id and leed_id corresponding to partner_meter_id"""
def get_meter_id(leed_meter_id):
	building_leed_id = "SELECT Leed_Id FROM %s WHERE id = (SELECT property_id FROM %s WHERE Energystar_meter_Id='%s') "\
	 % (energystar_property, energystar_meters, leed_meter_id)
	
	get_db = connect_db()
	leed_db = get_db[0]
	db_cursor = get_db[1]
	db_cursor.execute(building_leed_id)
	leed_building = db_cursor.fetchone()
	leed_db.close()
	
	# Call the MeterCollection API to get the id corresponding to partner meter id
	leedon_building_meters = "/buildings/LEED:%s/meters/" %(leed_building[0])
	 
	leedon_meters_url = urljoin(leedon_url, leedon_building_meters)
	print leedon_meters_url,'-----leedon_meters_url-----'
	try:
		leedid_meters = requests.get(leedon_meters_url, auth=leed_auth)
		list_leedid_meters = leedid_meters.json()

		# From list of meters, see if the meter exists in the leedon and get its id and return response according to it
		if list_leedid_meters:
			if isinstance(list_leedid_meters, list):
				for leed in list_leedid_meters:
					if leed['partner_meter_id'] == leed_meter_id:
						leedon_meter_id = leed['id']
						# print leedon_meter_id
						flag = 1
						return [leedon_meter_id, leed_building[0], flag]

				flag = 0
				return flag
		else:
			flag = 0
			return flag

	except Exception as e:
		print "---Exception in get_meter_id()---", e
		flag = 0
		return flag


#######################################################################################################################

"""Post the data to leedon using the MeterConsumption API whenever there are changes"""
def add_to_leedon(*argz):
	headers = {'content-type': 'application/json'}

	leed_flag = 0
	# Leed Building Id corresponding to the meter id
	check_leed = get_meter_id(argz[0])
	if isinstance(check_leed, list):
		leedon_id = check_leed[0]
		leedon_building = check_leed[1]
		leed_flag = check_leed[2]
	else:
		leed_flag = check_leed

	if leed_flag:
		pass
	else:
		print "Meter Id not added by leedon"
		return False

	# Call the MeterConsumption API to post the updated values
	leedon_post_url = "/buildings/LEED:%s/meters/ID:%s/data/" %(leedon_building, leedon_id)
	leed_consumption_url = combine_url(leedon_url, leedon_post_url)
	# print "Post data to", leed_consumption_url

	inc_end_date = (datetime.strptime(argz[2], FMD) + timedelta(days=1)).strftime(FMD)
	leed_data = {"start_date": argz[1], "end_date": inc_end_date, "reading": argz[3],"partner_consumption_id": argz[5]}
	# print "Data to leedon", leed_data
	energystar_timestamp = argz[4]
	
	try:
		post_data = requests.post(leed_consumption_url, data=dumps([leed_data]), auth=leed_auth, headers=headers,)
		received_data = post_data
		print "Post Meter_ID", argz[0],"Post data to", leed_consumption_url,"Data to leedon", leed_data,\
		'---Response to Post---', received_data.text
		# leedon_updates.append(argz[4])

		# To check if there is an Overlapping of data of start_date and end_date
		check_post = received_data.json()

		# No overlapping of data while posting the data to leedon
		if check_post[0]['result'] == "success":
			leedon_get = requests.get(leed_consumption_url, auth=leed_auth)
			# print leedon_get.text
			leedon_json = leedon_get.json()
			leedon_timestamp = leedon_json[0]['updated_at']
			
			leedon_timestamp_update = dateparser.parse(str(leedon_timestamp)[:19])
			timestamp_diff = leedon_timestamp_update - energystar_timestamp
			time_diff = (datetime.strptime("00:00:00", FMT) + timestamp_diff).strftime(FMT)
			print "energystar_timestamp", energystar_timestamp.strftime(FMF),"leedon_timestamp", \
			leedon_timestamp_update.strftime(FMF),"Timestamp difference", time_diff
			
			return leedon_building

		# Increment Start date by 1 day
		elif check_post[0]['result'] == "failure":
			# Get the Meter Name corresponding to the Meter ID
			get_db = connect_db()
			leed_db = get_db[0]
			db_cursor = get_db[1]
			db_cursor.execute(("SELECT Meter_Name FROM %s WHERE Energystar_meter_Id = '%s' ")\
			 %(energystar_meters, argz[0]))
			post_meter = db_cursor.fetchone()
			leed_db.close()

			"""For checking the overlapping data from leedon for the start_date and end_date by comparision"""
			# inc_end_date = (datetime.strptime(argz[2], FMD) + timedelta(days=1)).strftime(FMD)
 			leed_range_date = "/buildings/LEED:%s/meters/ID:%s/data/?start=%s&end=%s&order=recent" %(leedon_building,\
			 leedon_id, argz[1], inc_end_date)
			range_date_url = urljoin(leedon_url, leed_range_date)

			# Get the data for that period
			leedon_get_dates = requests.get(range_date_url, auth=leed_auth)
			leedon_dates_json = leedon_get_dates.json()
			
			if leedon_dates_json:
				print "Get data for overlapping dates--", leedon_get_dates.text, "at", range_date_url
				for d in leedon_dates_json:
					leedon_start_date = str(d['start_date'])[:10]
					leedon_end_date = str(d['end_date'])[:10]
					leedon_reading = d['reading']
					leedon_consumption_id = d['partner_consumption_id']

					# If same start_date and end_date
					if leedon_start_date == str(argz[1]) and leedon_end_date == inc_end_date:
						if leedon_reading == float(argz[3]) and leedon_consumption_id == int(argz[5]):
							print "Reading already exists for the same dates"
							return False

						else:
							# Reading not same, but dates are same
							leedon_delete_url = "/buildings/LEED:%s/meters/ID:%s/data/?start=%s&end=%s" \
							%(leedon_building, leedon_id, leedon_start_date, leedon_end_date)
							leedon_delete_date_url = urljoin(leedon_url, leedon_delete_url)
							# Delete the data from leedon for that dates
							requests.delete(leedon_delete_date_url, auth=leed_auth)

							# Post the data to leedon
							post_data = requests.post(leed_consumption_url, data=dumps([leed_data]), auth=leed_auth,\
							 headers=headers,)
							print "Re-posting same data after deletion", post_data.text
							re_post_data = post_data.json()
							if re_post_data[0]['result'] == "success":
								# ws2.append([leedon_building, post_meter[0], argz[1], argz[2], argz[3], argz[1],\
								#  argz[2]])
								# wb.save(filename = leedon_excel)
								return leedon_building
							else:
								pass
					else:
						# Start_dates and end_Date not same
						pass
			else:
				# No data for the called api in between the start_date and end_Date
				pass

			# If start_end dates don't match, then there must be some problem with the data
			check_leed_data = dict(leed_data)
			new_start_date = check_leed_data['start_date']
			# Increment start_date
			check_leed_data['start_date'] = (datetime.strptime(new_start_date, FMD) + timedelta(days=1))\
			.strftime(FMD)
			post_data = requests.post(leed_consumption_url, data=dumps([check_leed_data]), auth=leed_auth,\
			 headers=headers,)
			received_data = post_data.json()
		
		print '---Response to Post After increment start date---', post_data.text
		# No overlapping of data after increment of start date by 1 day and add to sheet
		if received_data[0]['result'] == "success":
			ws2.append([leedon_building, post_meter[0], argz[1], inc_end_date, argz[3], check_leed_data['start_date'],\
			 inc_end_date])
		 	wb.save(filename = leedon_excel)

			leedon_get = requests.get(leed_consumption_url, auth=leed_auth)
			#print leedon_get.text
			leedon_json = leedon_get.json()
			leedon_timestamp = leedon_json[0]['updated_at']
			
			leedon_timestamp_update = dateparser.parse(str(leedon_timestamp)[:19])
			timestamp_diff = leedon_timestamp_update - energystar_timestamp
			time_diff = (datetime.strptime("00:00:00", FMT) + timestamp_diff).strftime(FMT)
			print "energystar_timestamp", energystar_timestamp.strftime(FMF),"leedon_timestamp", \
			leedon_timestamp_update.strftime(FMF),"Timestamp difference", time_diff
			
			return leedon_building

		# Decrement End_Date by 1 day
		elif received_data[0]['result'] == "failure":
			check_leed_data = dict(leed_data)
			new_end_date = check_leed_data['end_date']
			# Decrement end_date
			check_leed_data['end_date'] = (datetime.strptime(new_end_date, FMD) - timedelta(days=1)).strftime(FMD)
			post_data = requests.post(leed_consumption_url, data=dumps([check_leed_data]), auth=leed_auth,\
			 headers=headers,)
			received_data1 = post_data.json()

		print '---Response to Post After decrement end date---', post_data.text 
		# No overlapping of data after decrement of end date by 1 and sheet
		if received_data1[0]['result'] == "success":
			ws2.append([leedon_building, post_meter[0], argz[1], inc_end_date, argz[3], argz[1],\
			 check_leed_data['end_date']])
		 	wb.save(filename = leedon_excel)

			leedon_get = requests.get(leed_consumption_url, auth=leed_auth)
			#print leedon_get.text
			leedon_json = leedon_get.json()
			leedon_timestamp = leedon_json[0]['updated_at']
			
			leedon_timestamp_update = dateparser.parse(str(leedon_timestamp)[:19])
			timestamp_diff = leedon_timestamp_update - energystar_timestamp
			time_diff = (datetime.strptime("00:00:00", FMT) + timestamp_diff).strftime(FMT)
			print "energystar_timestamp", energystar_timestamp.strftime(FMF),"leedon_timestamp", \
			leedon_timestamp_update.strftime(FMF),"Timestamp difference", time_diff
			
			return leedon_building
		
		# Increment Start_date by 1 day as well as decrement End_date by 1 day
		elif received_data1[0]['result'] == "failure":
			check_leed_data = dict(leed_data)
			new_start_date = check_leed_data['start_date']
			new_end_date = check_leed_data['end_date']
			check_leed_data['start_date'] = (datetime.strptime(new_start_date, FMD) + timedelta(days=1))\
			.strftime(FMD)
			check_leed_data['end_date'] = (datetime.strptime(new_end_date, FMD) - timedelta(days=1)).strftime(FMD)
			post_data = requests.post(leed_consumption_url, data=dumps([check_leed_data]), auth=leed_auth,\
			 headers=headers,)
			received_data2 = post_data.json()
		
		print '---Response to Post After both increment and decrement dates---', post_data.text 
		# No overlapping after changing both dates
		if received_data2[0]['result'] == "success":
			ws2.append([leedon_building, post_meter[0], argz[1], inc_end_date, argz[3], check_leed_data['start_date'],\
			 check_leed_data['end_date']])
		 	wb.save(filename = leedon_excel)

			leedon_get = requests.get(leed_consumption_url, auth=leed_auth)
			#print leedon_get.text
			leedon_json = leedon_get.json()
			leedon_timestamp = leedon_json[0]['updated_at']
			
			leedon_timestamp_update = dateparser.parse(str(leedon_timestamp)[:19])
			timestamp_diff = leedon_timestamp_update - energystar_timestamp
			time_diff = (datetime.strptime("00:00:00", FMT) + timestamp_diff).strftime(FMT)
			print "energystar_timestamp", energystar_timestamp.strftime(FMF),"leedon_timestamp", \
			leedon_timestamp_update.strftime(FMF),"Timestamp difference", time_diff
			
			return leedon_building

		elif received_data2[0]['result'] == "failure":
			# Add to Overlapping Meter Data Sheet even after making changes
			ws1.append([leedon_building, post_meter[0], argz[1], inc_end_date, argz[3]])
		 	wb.save(filename = leedon_excel)
			return False

	except Exception as e:
		print "---Exception in add_to_leedon()---", e
		return False

#######################################################################################################################

"""Function to check whether the consumption_id gets deleted or not from energystar"""
def check_delete_data(list_cons_id, mid):
	energystar_consid = []

	get_db = connect_db()
	leed_db = get_db[0]
	db_cursor = get_db[1]

	get_db_cons_id = "SELECT Energystar_consumption_Id FROM %s WHERE meter_id = (SELECT id FROM %s WHERE\
	 Energystar_meter_Id = '%s') ORDER BY Start_Date DESC LIMIT 24" %(energystar_data, energystar_meters, mid)
	db_cursor.execute(get_db_cons_id)
	get_db_consumption_ids = db_cursor.fetchall()

	for x in get_db_consumption_ids:
		energystar_consid.append(int(x[0]))

	delete_score = 0
	date_to_be_deleted = []
	for l in energystar_consid:
		if l not in list_cons_id:
			delete_score += 1
			print "Delete consumption_id", l

			# Get the id of meter from leedon by comparing the partner meter id
			delete_leedon = get_meter_id(mid)
			if isinstance(delete_leedon, list):
				leedon_id = delete_leedon[0]
				leedon_building = delete_leedon[1]
			else:
				continue

			leedon_delete = "SELECT Start_Date, End_Date FROM %s WHERE Energystar_consumption_Id = '%s' " \
			%(energystar_data, l)
			db_cursor.execute(leedon_delete)
			leedon_delete_date = db_cursor.fetchone()
			
			leedon_delete_url = "/buildings/LEED:%s/meters/ID:%s/data/?start=%s&end=%s" %(leedon_building, leedon_id,\
			 leedon_delete_date[0], leedon_delete_date[1])
			leedon_delete_date_url = urljoin(leedon_url, leedon_delete_url)
			print "DELETE FROM %s WHERE Energystar_consumption_Id =",energystar_data,"Delete", leedon_delete_date_url

			# Store the list of Start_Date for the consumption_ids to be deleted
			date_to_be_deleted.append(leedon_delete_date[0])

			# For deletion from both db as well as leedon
			try:
				#requests.delete(leedon_delete_date_url, auth=leed_auth)
				#db_cursor.execute('DELETE FROM %s WHERE Energystar_consumption_Id = "%s" ' %(energystar_data, l))
				pass
			except Exception as e:
				print "---Exception in check_delete_data---", e

	leed_db.commit()
	leed_db.close()

	if delete_score:
		return [leedon_building, date_to_be_deleted]
	else:
		return False

#######################################################################################################################

"""Call the Recompute score api by passing the earliest_date from list of all updated start_dates of meter"""
def recompute_score(recompute_leed, recompute_start):
	recompute_current_date = datetime.now().strftime(FMD)
	recompute_past_date = (datetime.now() - relativedelta(months=12)).strftime(FMD)

	# Delete the dates if less than one year
	if isinstance(recompute_start, list):
		min_recompute_start = [x for x in recompute_start if (x >= recompute_past_date)]
		earliest_date = min(min_recompute_start)
	else:
		earliest_date = recompute_start

	# Get the number of months between earliest date from list of updated dates and the current date
	date_diff = relativedelta(datetime.strptime(recompute_current_date, FMD), datetime.strptime(earliest_date, FMD))
	if date_diff.months:
		no_of_months = date_diff.months
	else:
		no_of_months = 0

	try:
		if no_of_months:
			for x in range(no_of_months+1):
				at_date = (datetime.strptime(earliest_date, FMD) - relativedelta(day=1) \
					+ relativedelta(months=x)).strftime(FMD)
				recomp_url = "/buildings/LEED:%s/recompute_score/?at=%s" %(recompute_leed, at_date)
				recompute_url = urljoin(leedon_url, recomp_url)
				print "Recompute URL", recompute_url
				requests.get(recompute_url, auth=leed_auth)

			recomp_today_url = "/buildings/LEED:%s/recompute_score/?at=%s" %(recompute_leed, recompute_current_date)
			recompute_today_url = urljoin(leedon_url, recomp_today_url)
			print "Recompute URL", recompute_today_url
			requests.get(recompute_today_url, auth=leed_auth)
		
		else:
			at_date = (datetime.strptime(earliest_date, FMD) - relativedelta(day=1)).strftime(FMD)
			recomp_url = "/buildings/LEED:%s/recompute_score/?at=%s" %(recompute_leed, at_date)
			recompute_url = urljoin(leedon_url, recomp_url)
			print "Recompute URL", recompute_url
			requests.get(recompute_url, auth=leed_auth)

			if at_date != recompute_current_date:
				recomp_today_url = "/buildings/LEED:%s/recompute_score/?at=%s" %(recompute_leed,recompute_current_date)
				recompute_today_url = urljoin(leedon_url, recomp_today_url)
				requests.get(recompute_today_url, auth=leed_auth)
	
	except Exception as e:
		print "---Exception in recompute_score()---", e
	
	return
	'''
	# Call the Performance API to recalulate the score
	leedon_performance = "/buildings/LEED:%s/performance/" %(recompute_leed)
	performance_url = urljoin(leedon_url, leedon_performance)
	get_performance = requests.get(performance_url, auth=leed_auth)
	print get_performance.text
	'''

#######################################################################################################################

"""Saving the meter consumption data into MeterConsumption table according to their start date and end date"""
def leed_consumptions(*args):
	# Calculate the meterconsumption link using *args
	if args:
		# if isinstance(args[0], int):
		if len(args) == 1:
			meter_link = "/meter/%s/consumptionData" % args[0]
			consumption_link = urljoin(energystar_url, meter_link)
			print "\nCheck Updates at", consumption_link
	
		# elif isinstance(args[0], str):
		elif len(args) == 2:
			consumption_link = args[1]
			print consumption_link
	
	"""Getting the meterconsumption url"""
	try:
		get_consumption = urllib2.Request(consumption_link)
		get_consumption.add_header("Authorization", "Basic %s" %energy_auth)
		consumption = urllib2.urlopen(get_consumption).read()
		meterconsumption = xmltodict.parse(consumption)['meterData']
		# meterconsumption1 = loads(dumps(meterconsumption, sort_keys=True, indent=4, separators=(',', ': ')))
		current_date = datetime.now().strftime(FMD)

		# list_of_consumption_ids = []
		meterid = args[0]
		recompute_building = 0
		recompute_start_dates = []

		"""MeterConsumption saved in table according to their meter_id"""
		# try:
		if isinstance(meterconsumption['meterConsumption'], list):
			for mc in meterconsumption['meterConsumption']:
				if len(list_of_consumption_ids) < 24:

					get_ldb = connect_db()
					leed_ldb = get_ldb[0]
					ldb_cursor = get_ldb[1]
					
					# meterid = args[0]
					consumption_id = mc['id']
					list_of_consumption_ids.append(int(consumption_id))
					# print consumption_id

					lastupdateddate = mc['audit']['lastUpdatedDate']
					last_updated_timestamp = dateparser.parse(str(lastupdateddate)[:19])
					# last_updated_timestamp = datetime.strptime(str(lastupdateddate), '%Y-%m-%dT%H:%M:%S.%fZ')
					
					start_date = mc['startDate']
					end_date = mc['endDate']
					
					try:
						cost = mc['cost']
					except:
						cost = 0
					
					try:
						meter_usage = mc['usage']
					except:
						meter_usage = 0

					"""Update the value into the MeterConsumption Table if any change by checking its timestamp"""
					if ldb_cursor.execute(("SELECT (1) FROM %s WHERE Energystar_consumption_Id = '%s' AND\
					 (last_updated < '%s' OR last_updated IS NULL)") % (energystar_data, consumption_id,\
					  last_updated_timestamp)):
						
						meter_consumption = 'UPDATE %s SET Start_Date = "%s", End_Date = "%s", Cost = "%s", \
						meterUsage = "%s", Time_stamp = "%s", last_updated = "%s" WHERE \
						Energystar_consumption_Id = "%s" ' % (energystar_data, start_date, end_date, cost,\
						 meter_usage, current_date, last_updated_timestamp, consumption_id)
						
						ldb_cursor.execute(meter_consumption)
						print "Update to db and Post to Leedon", consumption_id, start_date, end_date, cost,\
						 meter_usage, current_date, last_updated_timestamp

						leed_ldb.commit()
						leed_ldb.close()
						# Call to the function to post the data to the leedon
						leed_recompute = add_to_leedon(meterid, start_date, end_date, meter_usage,\
						 last_updated_timestamp, consumption_id)
						
						if leed_recompute:
							# leedon_updates.append(consumption_id)
							recompute_building = leed_recompute
							recompute_start_dates.append(start_date)

					elif ldb_cursor.execute(("SELECT (1) FROM %s WHERE Energystar_consumption_Id = '%s' ") \
						% (energystar_data, consumption_id)):
						# print "skip"
						leed_ldb.close()
						continue

					else:
						meter_consumption = 'INSERT IGNORE INTO %s \
						(meter_id, Energystar_consumption_Id, Start_Date, End_Date, Cost, meterUsage, Time_stamp,\
						 last_updated) VALUES ((SELECT id FROM %s WHERE Energystar_meter_Id = "%s"), "%s","%s","%s",\
						 "%s","%s","%s","%s") ' % (energystar_data, energystar_meters, meterid, consumption_id,\
						  start_date, end_date, cost, meter_usage, current_date, last_updated_timestamp)

						ldb_cursor.execute(meter_consumption)
						leed_ldb.commit()
						leed_ldb.close()
						print "Insert to db and Post to Leedon",consumption_id, start_date, end_date, cost,\
						 meter_usage, current_date, last_updated_timestamp
						# Call to the function to post the data to the leedon
						leed_recompute = add_to_leedon(meterid, start_date, end_date, meter_usage,\
						 last_updated_timestamp, consumption_id)
						
						if leed_recompute:
							# leedon_updates.append(consumption_id)
							recompute_building = leed_recompute
							recompute_start_dates.append(start_date)

				else:
					# When the number of consumption_ids extends 24
					break

			# Recompute score in case of any updation
			if recompute_building:
				print "\nRecomputing the score"
				recompute_score(recompute_building, recompute_start_dates)

		# If the page involves the consumption data of a meter only
		elif isinstance(meterconsumption['meterConsumption'], dict):
			if len(list_of_consumption_ids) < 24:
				# meterid = args[0]
				consumption_id = meterconsumption['meterConsumption']['id']
				list_of_consumption_ids.append(int(consumption_id))
				# print consumption_id

				lastupdateddate = meterconsumption['meterConsumption']['audit']['lastUpdatedDate']
				last_updated_timestamp = dateparser.parse(str(lastupdateddate)[:19])
				start_date = meterconsumption['meterConsumption']['startDate']
				end_date = meterconsumption['meterConsumption']['endDate']
				
				try:
					cost = meterconsumption['meterConsumption']['cost']
				except:
					cost = 0
				
				try:
					meter_usage = meterconsumption['meterConsumption']['usage']
				except:
					meter_usage = 0

				get_ldb = connect_db()
				leed_ldb = get_ldb[0]
				ldb_cursor = get_ldb[1]

				"""Updating the value into the MeterConsumption Table"""
				if ldb_cursor.execute(("SELECT (1) FROM %s WHERE Energystar_consumption_Id = '%s' AND\
				 (last_updated < '%s' OR last_updated IS NULL)") % (energystar_data, consumption_id,\
				  last_updated_timestamp)):
					
					meter_consumption = 'UPDATE %s SET Start_Date = "%s", End_Date = "%s", Cost = "%s", \
					meterUsage = "%s", Time_stamp = "%s", last_updated = "%s" WHERE Energystar_consumption_Id = "%s" '\
					 % (energystar_data, start_date, end_date, cost, meter_usage, current_date,\
					  last_updated_timestamp, consumption_id)
					
					ldb_cursor.execute(meter_consumption)
					leed_ldb.commit()
					leed_ldb.close()
					print "Update to db and Post to Leedon", consumption_id, start_date, end_date, cost, meter_usage, \
					current_date, last_updated_timestamp
					# Call to the function to post the data to the leedon
					leed_recompute = add_to_leedon(meterid, start_date, end_date, meter_usage, last_updated_timestamp,\
					 consumption_id)
					if leed_recompute:
						# leedon_updates.append(consumption_id)
						recompute_building = leed_recompute
						recompute_start_dates.append(start_date)

				elif ldb_cursor.execute(("SELECT (1) FROM %s WHERE Energystar_consumption_Id = '%s' ") \
					% (energystar_data, consumption_id)):
					# print "skip"
					leed_ldb.close()
					pass

				else:
					meter_consumption = 'INSERT IGNORE INTO %s \
					(meter_id, Energystar_consumption_Id, Start_Date, End_Date, Cost, meterUsage, Time_stamp,\
					 last_updated) VALUES ((SELECT id FROM %s WHERE Energystar_meter_Id = "%s"), "%s","%s","%s","%s",\
					 "%s", "%s", "%s") ' % (energystar_data, energystar_meters, meterid, consumption_id, start_date,\
					  end_date, cost, meter_usage, current_date, last_updated_timestamp)
				
					ldb_cursor.execute(meter_consumption)
					leed_ldb.commit()
					leed_ldb.close()

					print "Insert to db and Post to Leedon",consumption_id, start_date, end_date, cost, meter_usage, \
					current_date, last_updated_timestamp
					leed_recompute = add_to_leedon(meterid, start_date, end_date, meter_usage, last_updated_timestamp,\
					 consumption_id)
					if leed_recompute:
						# leedon_updates.append(consumption_id)
						recompute_building = leed_recompute
						recompute_start_dates.append(start_date)
			
			# Recompute score in case of any updation
			if recompute_building:
				print "Recomputing the Score"
				recompute_score(recompute_building, recompute_start_dates)
	
	except Exception as e:
		print "---Exception in leed_consumptions()---", e
	
	# For getting the next page link of meter consumption data if it exists
	new_link = 0
	try:
		'''
		if isinstance(meterconsumption['links']['link'], list):
			for metcon in meterconsumption['links']['link']:	
				if metcon['@linkDescription'] == "next page":
					next_link = metcon['@link']
					next_page = urljoin(base_url, next_link)
					new_link = 1
		'''
		if isinstance(meterconsumption['links']['link'], dict):
			if meterconsumption['links']['link']['@linkDescription'] == "next page":
				next_link = meterconsumption['links']['link']['@link']
				next_page = urljoin(energystar_url, next_link)
				new_link = 1
			else:
				new_link = 0
	except:
		new_link = 0

	# If next page exists for the meterconsumption of a particular meter
	if new_link:
		leed_consumptions(args[0], next_page)

	# return leedon_updates
	return

#######################################################################################################################


"""Infinite loop that runs in form of loops showing the count, updates and timestamp"""
count = 0
log_date = datetime.now().strftime('%Y%m%d')
filename = 'logs/auto_sync_%s.txt' %(log_date)

while (1>0):
	# if count==2:
	#    break
	"""List of Energystar_consumption_Id updated or added by energystar"""
	# leedon_updates = []
	
	current_time = datetime.now()
	# To save the log to auto_sync_$dt.txt file depending on the date
	if current_time.strftime('%Y%m%d') != log_date:
		log_date = current_time.strftime('%Y%m%d')
		filename = 'logs/auto_sync_%s.txt' %(log_date)

	sys.stdout = open(filename, 'a')

	# print "\n\nLoop Count", count
	print count,"Loop Timestamp", current_time.strftime('%c')
	current_time = current_time.strftime(FMT)

	'''
	# For checking if the time reaches 6AM when running the infinite loop
	if current_time > "05:30:00" and current_time <= "06:00:00":
		if (datetime.strptime(current_time, FMT) + diff).strftime(FMT) < "06:00:00":
			pass
		else:
			print "\nTerminate the script"
			sys.exit()
	'''

	"""Get the List of Building Leed_Id with building_status activated or trail_version_status true"""
	try:
		# if (count%25 == 0):
		get_db = connect_db()
		leed_db = get_db[0]
		db_cursor = get_db[1]	
		
		# Leed_ids in db
		buildings_db = []
		# Leedid activated or trail_version_status true in leedon
		buildings_activated = []
		
		# List of all leed_ids from db and store in building_db
		list_of_leedids = "SELECT Leed_Id FROM %s WHERE Leed_Id IS NOT NULL AND Leed_Id != 0 ORDER BY Leed_Id" \
		%(energystar_property)
		db_cursor.execute(list_of_leedids)
		building_leedids = db_cursor.fetchall()
		
		for b in building_leedids:
			buildings_db.append(int(b[0]))

		print "No of leed_ids in db", len(buildings_db)

		maxn = 1000
		# Getting the Leed_ids from leedon in 100 increments activated one else break the loop
		for z in xrange(0,maxn,100):
			try:
				buildings_list_url = "/buildings/?start="+str(z)
				buildings_url = urljoin(leedon_url, buildings_list_url)
				# print buildings_url
				check_building = requests.get(buildings_url, auth=leed_auth)
				if check_building.ok:
					print '----check_building.json()---',check_building.json()
					building_data = check_building.json()
					
					if building_data:
						for x in building_data:
							if (x['building_status'] == "activated") or (x['trial_version_status']):
								buildings_activated.append(x['leed_id'])
					else:
						print "All the activated leedids added"
						break

			except Exception as e:
				print "Error when getting activated buildings", e
				pass

		print "No of activated buildings in leedon", len(buildings_activated)
		# Common leed_ids in the leedon activated and the energystar
		activated_buildings = set(buildings_db) & set(buildings_activated)
		print "No of activated buildings of energystar in leedon", len(activated_buildings)
		# print "Activated buildings of energystar in leedon", activated_buildings
		
		'''
		# For checking the leed_ids one by one, use the above loop(this not used anymore)
		for b in building_leedids:
			try:
				buildings_leed = "/buildings/LEED:%s/" %(b[0])
				buildings_url = urljoin(leedon_url, buildings_leed)
				check_building = requests.get(buildings_url, auth=leed_auth)
				# print buildings_url,check_building.ok
				if check_building.ok:                                
					building_data = check_building.json()
					#buildings_activated.append(b[0])
					#print "Activated Leed_id", b[0]
				if building_data['building_status']=="activated" or building_data['trial_version_status']=="true":
					buildings_activated.append(b[0])
					#print "Activated Leed_id", b[0]
				# else:
				# print check_building.text,b[0],'building not found'

			except Exception as e:
				print e,"error in buildings_activated",check_building.text,b[0]
				pass
		
		# activated_buildings.append(1000000412)
		print len(buildings_activated),'no of leed_id property in energystar',buildings_activated
		'''
		# List of all the meters for activated leed_ids
		buildings_meter = []

		if activated_buildings:
			for a in activated_buildings:
				# list_of_leed_meters = "SELECT l.Energystar_meter_Id FROM %s l INNER JOIN %s p ON l.property_id = \
				# p.property_id WHERE p.Leed_Id IS NOT NULL ORDER BY l.meter_id" \
				# %(energystar_meters, energystar_property)
				list_of_leed_meters = "SELECT Energystar_meter_Id FROM %s WHERE property_id IN (SELECT id FROM %s \
				WHERE Leed_Id = '%s')" %(energystar_meters, energystar_property, a)
				db_cursor.execute(list_of_leed_meters)
				leed_meters = db_cursor.fetchall()			

				for x in leed_meters:
					buildings_meter.append(int(x[0]))

			leed_db.close()

			"""Call to leed_consumptions() for activated buildings only"""
			for x in buildings_meter:
				#print x, "Meter"
				list_of_consumption_ids = []
				leed_consumptions(x)
				
				'''
				# Check if any data gets deleted or not from the energystar
				print "\nChecking if any deletion"
				leed_delete = check_delete_data(list_of_consumption_ids, x)
				if leed_delete:
					delete_buildings = leed_delete[0]
					delete_start_dates = leed_delete[1]
					for d in delete_start_dates:
						recompute_score(delete_buildings, d)
				'''
		else:
			leed_db.close()

	except Exception as e:
		print "---Exception while loop count---", e


	# Loop count increment
	count += 1

	"""Calculate the running time"""
	end_time = datetime.now().strftime(FMT)
	diff = datetime.strptime(end_time, FMT) - datetime.strptime(current_time, FMT)
	print "\nEnd time of loop", end_time
	print "Running time of loop", diff


"""Close the database db after commiting changes to it"""
#db.close()
